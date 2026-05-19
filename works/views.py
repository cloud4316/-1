import json
import os
import shutil
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate, logout
from django.db.models import Count, Q, Avg, Max, Sum, F
from django.http import JsonResponse, HttpResponseNotAllowed
from django.contrib.auth.models import User
from django.utils import timezone
from django.core.paginator import Paginator
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.core.cache import cache
from .models import (PracticalWork, Solution, UserProgress, UserSession, PageView,
                     CodeCheck, TheoryModule, TheoryLesson, LessonProgress,
                     Quiz, Question, AnswerChoice, QuizAttempt,
                     Notification, TeacherComment,
                     Achievement, UserAchievement, WorkHint, UserHintUnlock,
                     DeadlineExtension, Subject, Announcement,
                     CircuitDraft, CircuitSolution)
from .code_runner import run_python_code, run_java_code, run_cpp_code, run_javascript_code
from .ai_checker import AICodeChecker
from .forms import SolutionForm, RegistrationForm, FullNameLoginForm, generate_username
from .utils import fix_file_encoding, save_file_with_correct_encoding
from datetime import timedelta, datetime, date


def _notify_email(user, subject, body):
    """Отправить email-уведомление пользователю (fail-safe)."""
    if not getattr(user, 'email', None):
        return
    try:
        from django.core.mail import send_mail
        from django.conf import settings as _settings
        send_mail(
            subject=subject,
            message=body,
            from_email=getattr(_settings, 'DEFAULT_FROM_EMAIL', 'noreply@algorithmmaster.local'),
            recipient_list=[user.email],
            fail_silently=True,
        )
    except Exception:
        pass


# ── Достижения ────────────────────────────────────────────────────────────────

_ACHIEVEMENT_DEFS = [
    ('first_solve',     '⭐', 'Первый шаг',        'Сдай первое задание правильно',              10),
    ('five_solves',     '🔥', 'В ударе',            'Реши 5 заданий правильно',                   25),
    ('ten_solves',      '🏆', 'Практик',            'Реши 10 заданий правильно',                  50),
    ('theory_start',    '📚', 'Теоретик',           'Изучи первый урок теории',                   5),
    ('theory_ten',      '🧠', 'Знаток теории',      'Изучи 10 уроков теории',                     30),
    ('streak_3',        '⚡', 'Серия × 3',          'Заходи 3 дня подряд',                        15),
    ('streak_7',        '🚀', 'Недельная серия',    'Заходи 7 дней подряд',                       40),
    ('perfect_score',   '🎯', 'Перфекционист',      'Получи максимальный балл за задание',        20),
]

def _ensure_achievements():
    """Создаёт записи Achievement если их нет."""
    for key, icon, title, desc, xp in _ACHIEVEMENT_DEFS:
        Achievement.objects.get_or_create(
            key=key,
            defaults={'title': title, 'description': desc, 'icon': icon, 'xp_reward': xp}
        )

def check_achievements(user):
    """Проверяет и выдаёт новые достижения пользователю."""
    try:
        _ensure_achievements()
        progress = UserProgress.objects.filter(user=user).first()
        if not progress:
            return
        correct_count = Solution.objects.filter(student=user, status='correct').count()
        theory_count  = LessonProgress.objects.filter(user=user, completed=True).count()
        has_perfect   = Solution.objects.filter(
            student=user, status='correct'
        ).filter(score__gte=F('work__max_score')).exists()

        conditions = {
            'first_solve':   correct_count >= 1,
            'five_solves':   correct_count >= 5,
            'ten_solves':    correct_count >= 10,
            'theory_start':  theory_count  >= 1,
            'theory_ten':    theory_count  >= 10,
            'streak_3':      progress.streak_days >= 3,
            'streak_7':      progress.streak_days >= 7,
            'perfect_score': has_perfect,
        }
        existing = set(
            UserAchievement.objects.filter(user=user)
            .values_list('achievement__key', flat=True)
        )
        for key, met in conditions.items():
            if met and key not in existing:
                try:
                    ach = Achievement.objects.get(key=key)
                    UserAchievement.objects.create(user=user, achievement=ach)
                    Notification.send(
                        user=user,
                        n_type='info',
                        title=f'Новое достижение: {ach.icon} {ach.title}',
                        message=ach.description,
                    )
                except Achievement.DoesNotExist:
                    pass
    except Exception:
        pass  # Не ломаем основной поток


def _active_announcements(request):
    """Вернуть активные объявления для текущего предмета + общие."""
    now = timezone.now()
    subject_slug = request.session.get('subject_slug', 'python')
    return Announcement.objects.filter(
        is_active=True
    ).filter(
        Q(expires_at__isnull=True) | Q(expires_at__gt=now)
    ).filter(
        Q(subject__isnull=True) | Q(subject__slug=subject_slug)
    ).select_related('author', 'subject').order_by('-created_at')[:5]


def home(request):  # Публичная страница — @login_required не нужен
    subjects = Subject.objects.filter(is_active=True)
    announcements = _active_announcements(request) if request.user.is_authenticated else []

    if request.user.is_authenticated:
        progress, created = UserProgress.objects.get_or_create(user=request.user)
        total_works = PracticalWork.objects.filter(is_active=True).count()
        completed_works = Solution.objects.filter(
            student=request.user,
            status__in=['correct', 'partially_correct']
        ).values('work').distinct().count()
        total_score = Solution.objects.filter(student=request.user).aggregate(Sum('score'))['score__sum'] or 0
        progress.total_works = total_works
        progress.completed_works = completed_works
        progress.total_score = total_score
        progress.average_score = total_score / completed_works if completed_works > 0 else 0
        progress.save()
        recent_solutions = Solution.objects.filter(student=request.user).order_by('-submitted_at')[:3]
        context = {
            'progress': progress,
            'recent_solutions': recent_solutions,
            'total_works': total_works,
            'completed_works': completed_works,
            'highlight_works': PracticalWork.objects.filter(is_active=True).order_by('order')[:3],
            'subjects': subjects,
            'announcements': announcements,
            'current_subject_slug': request.session.get('subject_slug', 'python'),
        }
    else:
        context = {
            'highlight_works': PracticalWork.objects.filter(is_active=True).order_by('order')[:3],
            'subjects': subjects,
        }
    return render(request, 'works/home.html', context)


def register(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            d = form.cleaned_data
            username = generate_username(d['last_name'], d['first_name'])
            user = User.objects.create_user(
                username=username,
                password=d['password1'],
                first_name=d['first_name'],
                last_name=d['last_name'],
                is_active=False,   # ждёт подтверждения преподавателя
            )
            # Сохраняем группу в профиле
            progress = UserProgress.objects.create(user=user)
            if d.get('group'):
                progress.group = d['group']
                progress.save()
            messages.success(request, 'Заявка отправлена')
            return render(request, 'works/pending_approval.html', {
                'full_name': f"{d['last_name']} {d['first_name']}",
                'group': d.get('group', ''),
            })
    else:
        form = RegistrationForm()
    return render(request, 'works/register.html', {'form': form})


def custom_login(request):
    """Вход по «Фамилия Имя» + пароль.
    Поддерживает два формата:
    — «Иванов Иван» (Фамилия Имя) — для студентов
    — «admin admin» или просто «admin» — для преподавателя/суперюзера
    """
    if request.user.is_authenticated:
        return redirect('home')

    error = None
    form = FullNameLoginForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        full_name = form.cleaned_data['full_name'].strip()
        password  = form.cleaned_data['password']

        user = None

        # 1. Пробуем FullNameBackend (Фамилия Имя)
        from works.backends import FullNameBackend
        backend = FullNameBackend()
        user = backend._find_user(full_name)

        # 2. Fallback: пробуем по username (для admin и суперюзеров)
        if user is None:
            from django.contrib.auth.models import User as _User
            # Если ввели "admin admin" — пробуем первое слово как username
            username_try = full_name.split()[0] if full_name else full_name
            user = _User.objects.filter(username__iexact=username_try).first()
            # Или точное совпадение всей строки с username
            if user is None:
                user = _User.objects.filter(username__iexact=full_name).first()

        if user is None:
            error = 'Пользователь не найден. Проверьте фамилию и имя.'
        elif not user.check_password(password):
            error = 'Неверный пароль.'
        elif not user.is_active:
            error = 'PENDING'
        else:
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            next_url = request.GET.get('next', '')
            return redirect(next_url if next_url else 'home')

    return render(request, 'works/login.html', {
        'form': form,
        'error': error,
    })


# ── Панель преподавателя ────────────────────────────────────────────────────

def _staff_required(view_func):
    """Декоратор: только для is_staff пользователей."""
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect(f'/accounts/login/?next={request.path}')
        if not request.user.is_staff:
            messages.error(request, 'Доступ запрещён.')
            return redirect('home')
        return view_func(request, *args, **kwargs)
    return wrapper


@_staff_required
def admin_panel(request):
    """Панель преподавателя: управление регистрациями студентов."""
    pending  = User.objects.filter(is_active=False, is_staff=False).order_by('last_name', 'first_name')
    approved = User.objects.filter(is_active=True,  is_staff=False).order_by('last_name', 'first_name')

    # Прогресс для одобренных
    progress_map = {
        p.user_id: p
        for p in UserProgress.objects.filter(user__in=approved).select_related('user')
    }
    for u in approved:
        u.progress = progress_map.get(u.id)

    return render(request, 'works/admin_panel.html', {
        'pending':  pending,
        'approved': approved,
    })


@_staff_required
@require_http_methods(['POST'])
def approve_user(request, user_id):
    """Одобрить регистрацию студента."""
    user = get_object_or_404(User, id=user_id, is_staff=False)
    user.is_active = True
    user.save()
    UserProgress.objects.get_or_create(user=user)
    Notification.send(
        user=user, n_type='approved',
        title='Регистрация одобрена!',
        message='Преподаватель одобрил вашу регистрацию. Можете войти в систему.',
        link='/login/',
    )
    messages.success(request, f'Студент {user.last_name} {user.first_name} допущен.')
    _notify_email(
        user,
        'Добро пожаловать в AlgorithmMaster!',
        f'Здравствуйте, {user.first_name}!\n\nВаша регистрация одобрена. Можете войти на платформу.',
    )
    return redirect('admin_panel')


@_staff_required
@require_http_methods(['POST'])
def reject_user(request, user_id):
    """Отклонить и удалить заявку."""
    user = get_object_or_404(User, id=user_id, is_staff=False)
    name = f'{user.last_name} {user.first_name}'
    user.delete()
    messages.success(request, f'Заявка студента {name} отклонена.')
    return redirect('admin_panel')


@_staff_required
@require_http_methods(['POST'])
def deactivate_user(request, user_id):
    """Заблокировать студента."""
    user = get_object_or_404(User, id=user_id, is_staff=False)
    user.is_active = False
    user.save()
    messages.success(request, f'Студент {user.last_name} {user.first_name} заблокирован.')
    return redirect('admin_panel')


@login_required
# ИСПРАВЛЕНО: убран @cache_page — кэшировал данные одного пользователя для всех
# и ломал AJAX-ответы (кэшировал первый тип ответа)
def work_list(request):
	is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
	has_filters = any([
		request.GET.get('q'),
		request.GET.get('difficulty') and request.GET.get('difficulty') != 'all',
		request.GET.get('topic') and request.GET.get('topic') != 'all',
		request.GET.get('sort') and request.GET.get('sort') != 'order'
	])
	subject_slug = request.session.get('subject_slug', 'python')
	qs = PracticalWork.objects.filter(is_active=True).select_related('subject').prefetch_related('solution_set')
	try:
		current_subject = Subject.objects.get(slug=subject_slug)
		qs = qs.filter(subject=current_subject)
	except Subject.DoesNotExist:
		current_subject = None
	q = request.GET.get('q')
	if q:
		qs = qs.filter(Q(title__icontains=q) | Q(description__icontains=q))
	difficulty = request.GET.get('difficulty')
	if difficulty in {'easy', 'medium', 'hard'}:
		qs = qs.filter(difficulty=difficulty)
	topic = request.GET.get('topic')
	if topic and topic != 'all':
		qs = qs.filter(topic=topic)
	sort = request.GET.get('sort')
	if sort == 'difficulty':
		qs = qs.order_by('difficulty', 'order')
	elif sort == 'deadline':
		qs = qs.order_by('deadline', 'order')
	else:
		qs = qs.order_by('order')
	user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
	is_mobile = any(device in user_agent for device in ['mobile', 'android', 'iphone', 'ipad'])
	is_tablet = any(device in user_agent for device in ['tablet', 'ipad'])
	if is_mobile:
		items_per_page = 6
	elif is_tablet:
		items_per_page = 9
	else:
		items_per_page = 12
	requested_items = request.GET.get('items_per_page')
	if requested_items and requested_items.isdigit():
		items_per_page = min(int(requested_items), 24)
	paginator = Paginator(qs, items_per_page)
	page_number = request.GET.get('page')
	page_obj = paginator.get_page(page_number)
	progress, _ = UserProgress.objects.get_or_create(user=request.user)
	user_solutions_raw = Solution.objects.filter(
		student=request.user,
		work__in=page_obj.object_list
	).select_related('work').order_by('work', '-submitted_at')
	user_solutions = {}
	for sol in user_solutions_raw:
		if sol.work_id not in user_solutions:
			user_solutions[sol.work_id] = sol
	works = list(page_obj.object_list)
	now = timezone.now()
	for work in works:
		user_solution = user_solutions.get(work.id)
		if user_solution:
			work.user_status = user_solution.status
			work.user_score = user_solution.score
			work.last_submission = user_solution.submitted_at
		else:
			work.user_status = 'not_started'
			work.user_score = 0
			work.last_submission = None
		work.is_overdue = bool(work.deadline and work.deadline < now and work.user_status not in ['correct', 'partially_correct'])
	if not has_filters:
		cache_key = f"user_stats_{request.user.id}"
		user_stats = cache.get(cache_key)
		if user_stats is None:
			solved_count = Solution.objects.filter(student=request.user, status__in=['correct', 'partially_correct']).values('work').distinct().count()
			in_progress_count = Solution.objects.filter(student=request.user, status__in=['submitted', 'checking']).values('work').distinct().count()
			user_stats = {'solved_count': solved_count, 'in_progress_count': in_progress_count}
			cache.set(cache_key, user_stats, 300)
		else:
			solved_count = user_stats['solved_count']
			in_progress_count = user_stats['in_progress_count']
	else:
		solved_count = Solution.objects.filter(student=request.user, status__in=['correct', 'partially_correct']).values('work').distinct().count()
		in_progress_count = Solution.objects.filter(student=request.user, status__in=['submitted', 'checking']).values('work').distinct().count()
	total_count = qs.count()
	topic_key_to_name = dict(PracticalWork.TOPIC_CHOICES)
	topics = qs.values('topic').annotate(cnt=Count('id')).order_by('topic')
	topic_filters = [
		{'key': row['topic'], 'name': topic_key_to_name.get(row['topic'], row['topic']), 'count': row['cnt']}
		for row in topics
	]
	context = {
		'works': works,
		'progress': progress,
		'page_obj': page_obj,
		'q': q or '',
		'difficulty': difficulty or 'all',
		'topic': topic or 'all',
		'sort': sort or 'order',
		'solved_count': solved_count,
		'in_progress_count': in_progress_count,
		'total_count': total_count,
		'topic_filters': topic_filters,
		'has_filters': has_filters,
		'items_per_page': items_per_page,
		'is_mobile': is_mobile,
		'is_tablet': is_tablet,
		'current_subject': current_subject,
	}
	if is_ajax:
		ajax_context = {
			'works': works, 'page_obj': page_obj,
			'q': q or '', 'difficulty': difficulty or 'all',
			'topic': topic or 'all', 'sort': sort or 'order',
			'total_count': total_count,
		}
		from django.template.loader import render_to_string
		html_content = render_to_string('works/works_grid.html', ajax_context, request=request)
		pagination_html = render_to_string('works/pagination.html', ajax_context, request=request)
		return JsonResponse({
			'html': html_content,
			'pagination': pagination_html,
			'current_page': page_obj.number,
			'total_pages': page_obj.paginator.num_pages,
			'total_count': total_count,
		})
	return render(request, 'works/work_list.html', context)


@login_required
def work_detail(request, work_id):
	work = get_object_or_404(PracticalWork, id=work_id)
	user_solutions = Solution.objects.filter(student=request.user, work=work).order_by('-submitted_at')
	best_solution = user_solutions.filter(status__in=['correct', 'partially_correct']).first()
	if request.method == 'POST':
		form = SolutionForm(request.POST, request.FILES)
		if form.is_valid():
			solution = form.save(commit=False)
			solution.student = request.user
			solution.work = work
			solution.original_filename = solution.code_file.name
			solution.status = 'submitted'
			solution.attempt_number = Solution.objects.filter(
				student=request.user, work=work
			).count() + 1
			solution.save()
			try:
				file_path = solution.code_file.path
				corrected_content = fix_file_encoding(file_path)
				save_file_with_correct_encoding(file_path, corrected_content)
			except Exception as e:
				print(f"Ошибка исправления кодировки: {e}")
			messages.success(request, 'Решение отправлено')
			return redirect('solution_detail', solution_id=solution.id)
	else:
		form = SolutionForm()
	user_extension = None
	if request.user.is_authenticated and not request.user.is_staff:
		user_extension = DeadlineExtension.objects.filter(
			user=request.user, work=work
		).first()
	return render(request, 'works/work_detail.html', {
		'work': work, 'form': form, 'solutions': user_solutions,
		'best_solution': best_solution,
		'total_attempts': user_solutions.count(),
		'successful_attempts': user_solutions.filter(status__in=['correct', 'partially_correct']).count(),
		'now': timezone.now(),
		'user_extension': user_extension,
	})


@login_required
def submission(request, work_id):
	if request.method != 'POST':
		return HttpResponseNotAllowed(['POST'])
	work = get_object_or_404(PracticalWork, id=work_id)
	form = SolutionForm(request.POST, request.FILES)
	if form.is_valid():
		solution = form.save(commit=False)
		solution.student = request.user
		solution.work = work
		solution.original_filename = solution.code_file.name
		solution.status = 'submitted'
		solution.attempt_number = Solution.objects.filter(
			student=request.user, work=work
		).count() + 1
		solution.save()
		try:
			file_path = solution.code_file.path
			corrected_content = fix_file_encoding(file_path)
			save_file_with_correct_encoding(file_path, corrected_content)
		except Exception as e:
			print(f"Ошибка исправления кодировки: {e}")
		messages.success(request, 'Решение отправлено')
		return render(request, 'works/submission.html', {'work': work, 'submission': solution})
	else:
		messages.error(request, 'Ошибка при загрузке файла')
		return redirect('work_detail', work_id=work_id)


def _calculate_streak_days(user):
    """Вычисляет количество дней подряд с активностью пользователя."""
    today = timezone.now().date()
    streak = 0
    check_date = today
    while True:
        has_activity = Solution.objects.filter(student=user, submitted_at__date=check_date).exists()
        if not has_activity:
            if check_date == today:
                check_date -= timedelta(days=1)
                continue
            break
        streak += 1
        check_date -= timedelta(days=1)
    return streak


@login_required
def profile(request):
    user = request.user
    progress, created = UserProgress.objects.get_or_create(user=user)
    solutions = Solution.objects.filter(student=user)
    total_attempts = solutions.count()
    successful_attempts = solutions.filter(status__in=['correct', 'partially_correct']).count()
    recent_solutions = solutions.order_by('-submitted_at')[:10]
    total_works = PracticalWork.objects.filter(is_active=True).count()
    completed_works = solutions.filter(status__in=['correct', 'partially_correct']).values('work').distinct().count()
    total_score = solutions.aggregate(Sum('score'))['score__sum'] or 0
    progress.total_works = total_works
    progress.completed_works = completed_works
    progress.total_score = total_score
    progress.average_score = total_score / completed_works if completed_works > 0 else 0
    # ИСПРАВЛЕНО: streak_days вычисляется по реальным данным
    progress.streak_days = _calculate_streak_days(user)
    progress.save()
    raw_activity = []
    for i in range(6, -1, -1):
        day_dt = timezone.now() - timedelta(days=i)
        date_start = day_dt.replace(hour=0, minute=0, second=0, microsecond=0)
        date_end = day_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
        day_sessions = UserSession.objects.filter(user=user, start_time__gte=date_start, start_time__lte=date_end)
        total_seconds = sum(session.duration_seconds for session in day_sessions)
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        count = solutions.filter(submitted_at__date=day_dt.date()).count()
        raw_activity.append({'date': day_dt.strftime('%d.%m'), 'count': count, 'hours': hours, 'minutes': minutes, 'total_seconds': total_seconds})
    max_hours = max((d['hours'] for d in raw_activity), default=0)
    activity_data = []
    for d in raw_activity:
        height = 8 + int(d['hours'] * 92 / max_hours) if max_hours > 0 else 8
        activity_data.append({'date': d['date'], 'count': d['count'], 'hours': d['hours'], 'minutes': d['minutes'], 'height': height})
    topic_key_to_name = dict(PracticalWork.TOPIC_CHOICES)
    total_by_topic = PracticalWork.objects.filter(is_active=True).values('topic').annotate(cnt=Count('id'))
    total_map = {row['topic']: row['cnt'] for row in total_by_topic}
    solved_work_ids = Solution.objects.filter(student=user, status__in=['correct', 'partially_correct']).values_list('work_id', flat=True).distinct()
    solved_by_topic = PracticalWork.objects.filter(id__in=solved_work_ids).values('topic').annotate(cnt=Count('id'))
    solved_map = {row['topic']: row['cnt'] for row in solved_by_topic}
    topics_progress = []
    for topic_key, topic_name in topic_key_to_name.items():
        total = total_map.get(topic_key, 0)
        solved = solved_map.get(topic_key, 0)
        percent = round(solved * 100 / total) if total > 0 else 0
        topics_progress.append({'name': topic_name, 'progress': percent, 'solved': solved, 'total': total})
    user_achievements = UserAchievement.objects.filter(user=request.user).select_related('achievement').order_by('-earned_at')
    all_achievements = Achievement.objects.all()
    earned_keys = set(ua.achievement.key for ua in user_achievements)
    check_achievements(request.user)
    return render(request, 'works/profile.html', {
        'progress': progress,
        'total_attempts': total_attempts,
        'successful_attempts': successful_attempts,
        'success_rate': round((successful_attempts / total_attempts * 100)) if total_attempts > 0 else 0,
        'recent_solutions': recent_solutions,
        'topics_progress': topics_progress,
        'activity_data': activity_data,
        'user_full_name': f"{user.first_name} {user.last_name}".strip() or user.username,
        'user_joined': user.date_joined,
        'level_progress_percent': progress.level_progress,
        'user_achievements': user_achievements,
        'all_achievements': all_achievements,
        'earned_keys': earned_keys,
    })


@login_required
@require_http_methods(["GET"])
def get_session_time(request):
    try:
        session_key = request.session.session_key
        if not session_key:
            request.session.save()
            session_key = request.session.session_key
        active_session = UserSession.objects.filter(user=request.user, session_key=session_key, is_active=True).first()
        if not active_session:
            return JsonResponse({'hours': 0, 'minutes': 0, 'seconds': 0, 'total_seconds': 0, 'status': 'no_session'})
        today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
        today_sessions = UserSession.objects.filter(user=request.user, start_time__gte=today_start, start_time__lte=today_end)
        total_today_seconds = sum(session.duration_seconds for session in today_sessions)
        if active_session.is_active:
            total_today_seconds += int((timezone.now() - active_session.start_time).total_seconds())
        hours = total_today_seconds // 3600
        minutes = (total_today_seconds % 3600) // 60
        seconds = total_today_seconds % 60
        return JsonResponse({'hours': hours, 'minutes': minutes, 'seconds': seconds, 'total_seconds': total_today_seconds, 'status': 'success'})
    except Exception as e:
        return JsonResponse({'hours': 0, 'minutes': 0, 'seconds': 0, 'total_seconds': 0, 'status': 'error', 'error': str(e)})


@login_required
@require_http_methods(["GET"])
def get_activity_data(request):
    try:
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=6)
        session_key = request.session.session_key
        if not session_key:
            request.session.save()
            session_key = request.session.session_key
        active_session = UserSession.objects.filter(user=request.user, session_key=session_key, is_active=True).first()
        activity_data = []
        for i in range(7):
            current_date = start_date + timedelta(days=i)
            # ИСПРАВЛЕНО: timezone.datetime не существует; используем datetime из stdlib
            day_start = timezone.make_aware(datetime.combine(current_date, datetime.min.time()))
            day_end = timezone.make_aware(datetime.combine(current_date, datetime.max.time()))
            day_sessions = UserSession.objects.filter(user=request.user, start_time__gte=day_start, start_time__lte=day_end)
            total_seconds = sum(session.duration_seconds for session in day_sessions)
            count = day_sessions.count()
            if current_date == timezone.now().date() and active_session and active_session.is_active:
                total_seconds += int((timezone.now() - active_session.start_time).total_seconds())
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            height = min((total_seconds / (8 * 3600)) * 100, 100)
            activity_data.append({'date': current_date.strftime('%d.%m'), 'count': count, 'hours': hours, 'minutes': minutes, 'height': int(height), 'total_seconds': total_seconds})
        return JsonResponse({'activity_data': activity_data, 'status': 'success'})
    except Exception as e:
        return JsonResponse({'activity_data': [], 'status': 'error', 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def check_code(request, solution_id):
    try:
        solution = Solution.objects.get(id=solution_id, student=request.user)
        check_type = request.POST.get('check_type', 'auto')
        # ИСПРАВЛЕНО: CodeCheck создаётся только здесь и передаётся в check_solution,
        # чтобы не было двойного создания записи
        code_check = CodeCheck.objects.create(
            solution=solution, status='in_progress',
            check_type=check_type, created_at=timezone.now()
        )
        import threading
        def run_check():
            checker = AICodeChecker()
            checker.check_solution(solution_id, code_check_id=code_check.id)
        thread = threading.Thread(target=run_check)
        thread.daemon = True
        thread.start()
        return JsonResponse({'status': 'success', 'check_id': code_check.id, 'message': 'Проверка кода запущена'})
    except Solution.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Решение не найдено'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Ошибка: {str(e)}'})


@login_required
@require_http_methods(["GET"])
def get_check_status(request, check_id):
    try:
        code_check = CodeCheck.objects.get(id=check_id)
        if code_check.solution.student != request.user:
            return JsonResponse({'status': 'error', 'message': 'У вас нет доступа к этой проверке'})
        response_data = {
            'status': 'success', 'check_status': code_check.status,
            'score': code_check.score, 'feedback': code_check.feedback,
            'suggestions': code_check.suggestions, 'errors': code_check.errors, 'warnings': code_check.warnings,
        }
        if code_check.completed_at:
            response_data['completed_at'] = code_check.completed_at.isoformat()
        return JsonResponse(response_data)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Ошибка: {str(e)}'})


@login_required
@require_http_methods(["POST"])
def test_code_locally(request, work_id):
    try:
        work = get_object_or_404(PracticalWork, id=work_id)
        code_content = request.POST.get('code', '')
        if not code_content:
            return JsonResponse({'status': 'error', 'message': 'Код не предоставлен'})
        language = work.language.lower() if work.language else 'python'
        input_data = work.input_example or None
        expected_output = (work.output_example or '').strip()
        # ИСПРАВЛЕНО: раньше код записывался во временный файл и путь передавался
        # в run_python_code(file_path, work) — конфликт сигнатур с code_runner.
        # Теперь передаём строку с кодом напрямую в функции из code_runner.
        if language == 'python':
            result = run_python_code(code_content, input_data)
        elif language == 'java':
            result = run_java_code(code_content, input_data)
        elif language in ['cpp', 'c']:
            result = run_cpp_code(code_content, input_data)
        elif language == 'javascript':
            result = run_javascript_code(code_content, input_data)
        else:
            return JsonResponse({'status': 'error', 'message': f'Язык {language} не поддерживается'})
        if result['status'] == 'success':
            actual_output = result.get('output', '').strip()
            test_passed = (actual_output == expected_output) if expected_output else True
            return JsonResponse({
                'status': 'success', 'test_passed': test_passed,
                'input': input_data or '', 'expected_output': expected_output,
                'actual_output': actual_output, 'error_output': '',
                'execution_time': '< 1 сек',
                'message': 'Тест пройден!' if test_passed else 'Тест не пройден'
            })
        else:
            return JsonResponse({
                'status': 'error', 'test_passed': False,
                'message': result.get('error', 'Ошибка выполнения'),
                'error_output': result.get('error', ''),
            })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Ошибка при тестировании: {str(e)}'})


@login_required
def results(request):
	solutions = Solution.objects.filter(student=request.user).order_by('-submitted_at')[:20]
	stats = {
		'total': Solution.objects.filter(student=request.user).count(),
		'solved': Solution.objects.filter(student=request.user, status__in=['correct', 'partially_correct']).count(),
	}
	return render(request, 'works/results.html', {'stats': stats, 'last_submissions': solutions})


@login_required
def works_userprogress(request):
    # ИСПРАВЛЕНО: убраны захардкоженные заглушки — теперь реальные данные из БД
    topic_key_to_name = dict(PracticalWork.TOPIC_CHOICES)
    total_by_topic = PracticalWork.objects.filter(is_active=True).values('topic').annotate(cnt=Count('id'))
    total_map = {row['topic']: row['cnt'] for row in total_by_topic}
    solved_work_ids = list(Solution.objects.filter(
        student=request.user, status__in=['correct', 'partially_correct']
    ).values_list('work_id', flat=True).distinct())
    solved_by_topic = PracticalWork.objects.filter(id__in=solved_work_ids).values('topic').annotate(cnt=Count('id'))
    solved_map = {row['topic']: row['cnt'] for row in solved_by_topic}
    topics = []
    for topic_key, topic_name in topic_key_to_name.items():
        total = total_map.get(topic_key, 0)
        solved = solved_map.get(topic_key, 0)
        percent = round(solved * 100 / total) if total > 0 else 0
        topics.append({'name': topic_name, 'percent': percent, 'solved': solved, 'total': total})
    stats = {
        'total': PracticalWork.objects.filter(is_active=True).count(),
        'solved': len(solved_work_ids),
    }
    return render(request, 'works/works_userprogress.html', {'topics': topics, 'stats': stats})


@login_required
def solution_detail(request, solution_id):
	solution = get_object_or_404(Solution, id=solution_id)
	if solution.student_id != request.user.id and not request.user.is_staff:
		return redirect('home')
	return render(request, 'works/solution_detail.html', {'solution': solution})


# ══════════════════════════════════════════════════════════════════════════════
# ТЕОРИЯ
# ══════════════════════════════════════════════════════════════════════════════

@login_required
def theory_list(request):
    subject_slug = request.session.get('subject_slug', 'python')
    try:
        current_subject = Subject.objects.get(slug=subject_slug)
    except Subject.DoesNotExist:
        current_subject = None

    qs = TheoryModule.objects.filter(is_active=True).prefetch_related('lessons', 'quizzes')
    if current_subject:
        qs = qs.filter(subject=current_subject)
    modules = list(qs)

    completed_lesson_ids = set(
        LessonProgress.objects.filter(user=request.user, completed=True)
        .values_list('lesson_id', flat=True)
    )

    unlock_status = get_module_unlock_status(request.user, modules)

    for module in modules:
        total = module.lessons.count()
        done  = sum(1 for l in module.lessons.all() if l.id in completed_lesson_ids)
        module.progress_done  = done
        module.progress_total = total
        module.progress_pct   = round(done * 100 / total) if total else 0
        module.is_completed   = (total > 0 and done == total)
        module.unlock         = unlock_status.get(module.id, {'theory': True, 'quiz': True, 'locked': False})

    return render(request, 'works/theory_list.html', {
        'modules': modules,
        'current_subject': current_subject,
        'subjects': Subject.objects.filter(is_active=True),
    })


@login_required
def theory_lesson(request, lesson_id):
    lesson = get_object_or_404(TheoryLesson, id=lesson_id)
    module = lesson.module

    # Все уроки модуля для навигации
    all_lessons = list(module.lessons.all())
    current_index = next((i for i, l in enumerate(all_lessons) if l.id == lesson_id), 0)
    prev_lesson = all_lessons[current_index - 1] if current_index > 0 else None
    next_lesson = all_lessons[current_index + 1] if current_index < len(all_lessons) - 1 else None

    # Отметка о прочтении + конспект
    progress_obj = LessonProgress.objects.filter(user=request.user, lesson=lesson).first()
    is_completed = bool(progress_obj and progress_obj.completed)
    user_notes   = progress_obj.notes if progress_obj else ''

    # Тесты для этого модуля
    module_quizzes = module.quizzes.filter(is_active=True)

    completed_ids = set(LessonProgress.objects.filter(
        user=request.user, completed=True
    ).values_list('lesson_id', flat=True))
    done_count = sum(1 for l in all_lessons if l.id in completed_ids)
    total_count = len(all_lessons)
    progress_pct = round(done_count * 100 / total_count) if total_count else 0

    return render(request, 'works/theory_lesson.html', {
        'lesson': lesson,
        'module': module,
        'all_lessons': all_lessons,
        'prev_lesson': prev_lesson,
        'next_lesson': next_lesson,
        'is_completed': is_completed,
        'current_index': current_index,
        'module_quizzes': module_quizzes,
        'completed_ids': completed_ids,
        'done_count': done_count,
        'total_count': total_count,
        'progress_pct': progress_pct,
        'user_notes': user_notes,
    })


@login_required
@require_http_methods(["POST"])
def mark_lesson_done(request, lesson_id):
    lesson = get_object_or_404(TheoryLesson, id=lesson_id)
    obj, created = LessonProgress.objects.get_or_create(
        user=request.user, lesson=lesson,
        defaults={'completed': True}
    )
    if not obj.completed:
        obj.completed = True
        obj.save()
    check_achievements(request.user)
    return JsonResponse({'status': 'ok', 'lesson_id': lesson_id})


# ── Конспекты ────────────────────────────────────────────────────────────────

@login_required
@require_http_methods(["POST"])
def save_notes(request, lesson_id):
    lesson = get_object_or_404(TheoryLesson, id=lesson_id)
    try:
        data = json.loads(request.body)
        notes_text = data.get('notes', '')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'status': 'error', 'message': 'bad json'}, status=400)

    obj, _ = LessonProgress.objects.get_or_create(
        user=request.user, lesson=lesson,
        defaults={'completed': False}
    )
    obj.notes = notes_text
    obj.save(update_fields=['notes', 'updated_at'])
    return JsonResponse({'status': 'ok', 'saved_at': obj.updated_at.strftime('%H:%M:%S')})


# ── Переключение предмета ─────────────────────────────────────────────────────

def switch_subject(request, slug):
    subject = get_object_or_404(Subject, slug=slug, is_active=True)
    request.session['subject_slug'] = subject.slug
    next_url = request.GET.get('next') or request.META.get('HTTP_REFERER', '/')
    return redirect(next_url)


# ── Объявления ────────────────────────────────────────────────────────────────

@login_required
def create_announcement(request):
    if not request.user.is_staff:
        return redirect('home')
    if request.method == 'POST':
        title      = request.POST.get('title', '').strip()
        body       = request.POST.get('body', '').strip()
        subject_id = request.POST.get('subject_id') or None
        expires    = request.POST.get('expires_at') or None
        if title:
            subj = Subject.objects.filter(id=subject_id).first() if subject_id else None
            exp  = None
            if expires:
                from django.utils.dateparse import parse_datetime
                exp = parse_datetime(expires)
            Announcement.objects.create(
                author=request.user, subject=subj,
                title=title, body=body,
                expires_at=exp,
            )
            messages.success(request, 'Объявление опубликовано.')
        return redirect(request.POST.get('next', 'home'))
    subjects = Subject.objects.filter(is_active=True)
    return render(request, 'works/announcement_form.html', {'subjects': subjects})


@login_required
@require_http_methods(["POST"])
def deactivate_announcement(request, pk):
    if not request.user.is_staff:
        return JsonResponse({'status': 'forbidden'}, status=403)
    ann = get_object_or_404(Announcement, pk=pk)
    ann.is_active = False
    ann.save(update_fields=['is_active'])
    return JsonResponse({'status': 'ok'})


@login_required
def announcement_list(request):
    if not request.user.is_staff:
        return redirect('home')
    qs = Announcement.objects.select_related('author', 'subject').order_by('-created_at')
    subject_filter = request.GET.get('subject')
    if subject_filter:
        qs = qs.filter(subject__slug=subject_filter)
    active_filter = request.GET.get('active')
    if active_filter == '1':
        qs = qs.filter(is_active=True)
    elif active_filter == '0':
        qs = qs.filter(is_active=False)
    announcements = list(qs[:50])
    now = timezone.now()
    for ann in announcements:
        ann.expired = bool(ann.expires_at and now > ann.expires_at)
    return render(request, 'works/announcement_list.html', {
        'announcements': announcements,
        'subjects': Subject.objects.filter(is_active=True),
        'subject_filter': subject_filter or '',
        'active_filter': active_filter or '',
    })


# ══════════════════════════════════════════════════════════════════════════════
# ТЕСТЫ
# ══════════════════════════════════════════════════════════════════════════════

@login_required
def quiz_list(request):
    subject_slug = request.session.get('subject_slug', 'python')
    qs = Quiz.objects.filter(is_active=True).select_related('module', 'module__subject').prefetch_related('questions')
    try:
        current_subject = Subject.objects.get(slug=subject_slug)
        qs = qs.filter(module__subject=current_subject)
    except Subject.DoesNotExist:
        current_subject = None
    quizzes = list(qs)

    attempt_map = {}
    for attempt in QuizAttempt.objects.filter(user=request.user).order_by('-score'):
        if attempt.quiz_id not in attempt_map:
            attempt_map[attempt.quiz_id] = attempt

    for quiz in quizzes:
        quiz.best_attempt = attempt_map.get(quiz.id)

    return render(request, 'works/quiz_list.html', {
        'quizzes': quizzes,
        'current_subject': current_subject,
    })


@login_required
def quiz_detail(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, is_active=True)
    questions = quiz.questions.prefetch_related('choices').all()

    best_attempt = QuizAttempt.objects.filter(
        user=request.user, quiz=quiz
    ).order_by('-score').first()

    return render(request, 'works/quiz.html', {
        'quiz': quiz,
        'questions': questions,
        'best_attempt': best_attempt,
    })


@login_required
@require_http_methods(["POST"])
def submit_quiz(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, is_active=True)
    questions = quiz.questions.prefetch_related('choices').all()

    total = questions.count()
    correct = 0
    results = {}

    for question in questions:
        correct_ids = set(
            question.choices.filter(is_correct=True).values_list('id', flat=True)
        )
        if question.q_type == 'single':
            chosen_raw = request.POST.get(f'q_{question.id}')
            chosen_ids = {int(chosen_raw)} if chosen_raw else set()
        else:
            chosen_ids = set(
                int(v) for v in request.POST.getlist(f'q_{question.id}')
            )

        is_correct = (chosen_ids == correct_ids)
        if is_correct:
            correct += 1

        results[str(question.id)] = {
            'chosen': list(chosen_ids),
            'correct': list(correct_ids),
            'is_correct': is_correct,
            'explanation': question.explanation,
        }

    score = round(correct * 100 / total) if total else 0
    passed = score >= quiz.pass_score

    attempt = QuizAttempt.objects.create(
        user=request.user,
        quiz=quiz,
        score=score,
        passed=passed,
        answers=results,
    )
    return render(request, 'works/quiz_result.html', {
        'quiz': quiz,
        'attempt': attempt,
        'questions': questions,
        'results': results,
        'results_json': json.dumps(results),
        'correct': correct,
        'errors': total - correct,
        'total': total,
        'score': score,
        'passed': passed,
    })


@login_required
@require_http_methods(["POST"])
def run_code_snippet(request):
    """Запуск произвольного кода из теоретических уроков (без привязки к задаче)."""
    code = request.POST.get("code", "").strip()
    if not code:
        return JsonResponse({"status": "error", "output": "Нет кода для выполнения"})
    try:
        result = run_python_code(code, input_data=None)
        output = result.get("output", "") or result.get("error", "")
        return JsonResponse({
            "status": result.get("status", "error"),
            "output": output or "(нет вывода)",
        })
    except Exception as e:
        return JsonResponse({"status": "error", "output": f"Ошибка сервера: {e}"})


# ══════════════════════════════════════════════════════════════════════════════
# УВЕДОМЛЕНИЯ
# ══════════════════════════════════════════════════════════════════════════════

@login_required
def notifications_view(request):
    notifs = Notification.objects.filter(user=request.user)
    # Помечаем все прочитанными при открытии страницы
    notifs.filter(is_read=False).update(is_read=True)
    return render(request, 'works/notifications.html', {'notifications': notifs})


@login_required
@require_http_methods(['POST'])
def mark_notification_read(request, notif_id):
    Notification.objects.filter(id=notif_id, user=request.user).update(is_read=True)
    return JsonResponse({'ok': True})


@login_required
@require_http_methods(['GET'])
def notifications_count(request):
    count = Notification.objects.filter(user=request.user, is_read=False).count()
    return JsonResponse({'count': count})


# ══════════════════════════════════════════════════════════════════════════════
# ЖУРНАЛ УСПЕВАЕМОСТИ (преподаватель)
# ══════════════════════════════════════════════════════════════════════════════

@_staff_required
def gradebook(request):
    group_filter = request.GET.get('group', '')
    students_qs  = User.objects.filter(is_active=True, is_staff=False).order_by('last_name', 'first_name')

    # Фильтрация по группе
    all_groups = list(
        UserProgress.objects.filter(group__gt='')
        .values_list('group', flat=True).distinct().order_by('group')
    )
    if group_filter:
        students_qs = students_qs.filter(userprogress__group=group_filter)

    students  = list(students_qs.select_related('userprogress'))
    works     = list(PracticalWork.objects.filter(is_active=True).order_by('order'))
    total_works_count = len(works)

    # Матрица оценок: {student_id: {work_id: best_score}}
    all_solutions = Solution.objects.filter(
        student__in=students,
        work__in=works,
    ).values('student_id', 'work_id', 'score', 'status').order_by('student_id', 'work_id', '-score')

    matrix = {}
    for sol in all_solutions:
        sid, wid = sol['student_id'], sol['work_id']
        if sid not in matrix:
            matrix[sid] = {}
        if wid not in matrix[sid]:
            matrix[sid][wid] = {'score': sol['score'], 'status': sol['status']}

    # Добавляем progress к студентам
    prog_map = {p.user_id: p for p in UserProgress.objects.filter(user__in=students)}
    for s in students:
        s.progress = prog_map.get(s.id)
        s.row = [matrix.get(s.id, {}).get(w.id) for w in works]
        scores = [c['score'] for c in matrix.get(s.id, {}).values() if c['score']]
        s.avg_score = round(sum(scores) / len(scores), 1) if scores else 0
        s.done_count = sum(1 for w in works if matrix.get(s.id, {}).get(w.id, {}).get('status') in ('correct','partially_correct'))

    export = request.GET.get('export')
    if export == 'csv':
        import csv
        from django.http import HttpResponse as HR
        resp = HR(content_type='text/csv; charset=utf-8-sig')
        resp['Content-Disposition'] = 'attachment; filename="gradebook.csv"'
        writer = csv.writer(resp)
        header = ['Студент', 'Группа'] + [f'ПР{w.order}' for w in works] + ['Среднее', 'Сдано']
        writer.writerow(header)
        for s in students:
            row_scores = []
            for w in works:
                cell = matrix.get(s.id, {}).get(w.id)
                row_scores.append(cell['score'] if cell else '')
            writer.writerow([
                f'{s.last_name} {s.first_name}',
                getattr(s.progress, 'group', ''),
                *row_scores,
                s.avg_score,
                s.done_count,
            ])
        return resp

    return render(request, 'works/gradebook.html', {
        'students': students,
        'works': works,
        'all_groups': all_groups,
        'group_filter': group_filter,
    })


# ══════════════════════════════════════════════════════════════════════════════
# РУЧНАЯ ПРОВЕРКА ПРЕПОДАВАТЕЛЕМ
# ══════════════════════════════════════════════════════════════════════════════

@_staff_required
def teacher_solutions(request):
    """Все решения с фильтрами для преподавателя."""
    qs = Solution.objects.select_related('student', 'work').prefetch_related('student__userprogress', 'teacher_comments').order_by('-submitted_at')
    group   = request.GET.get('group', '')
    status  = request.GET.get('status', '')
    work_id = request.GET.get('work', '')
    if group:
        qs = qs.filter(student__userprogress__group=group)
    if status:
        qs = qs.filter(status=status)
    if work_id:
        qs = qs.filter(work_id=work_id)

    from django.core.paginator import Paginator
    paginator = Paginator(qs, 30)
    page_obj  = paginator.get_page(request.GET.get('page'))

    all_groups = list(UserProgress.objects.filter(group__gt='')
                      .values_list('group', flat=True).distinct().order_by('group'))
    works_list = PracticalWork.objects.filter(is_active=True).order_by('order')

    pending_extensions = DeadlineExtension.objects.filter(status='pending').select_related('user', 'work').order_by('-requested_at')

    return render(request, 'works/teacher_solutions.html', {
        'page_obj': page_obj,
        'all_groups': all_groups,
        'works_list': works_list,
        'group': group,
        'status': status,
        'work_id': work_id,
        'pending_extensions': pending_extensions,
    })


@_staff_required
@require_http_methods(['POST'])
def manual_grade(request, solution_id):
    """Ручная оценка и комментарий от преподавателя."""
    solution = get_object_or_404(Solution, id=solution_id)
    score_raw = request.POST.get('score', '').strip()
    comment   = request.POST.get('comment', '').strip()
    new_status = request.POST.get('status', solution.status)

    score = None
    if score_raw.isdigit():
        score = min(int(score_raw), solution.work.max_score or 100)
        solution.score  = score
    solution.status = new_status
    solution.save()

    if comment:
        TeacherComment.objects.create(
            solution=solution,
            teacher=request.user,
            text=comment,
            score=score,
        )
        # Уведомляем студента
        Notification.send(
            user=solution.student,
            n_type='commented',
            title=f'Комментарий к «{solution.work.title}»',
            message=comment[:200],
            link=f'/solution/{solution.id}/',
        )

    if score is not None:
        Notification.send(
            user=solution.student,
            n_type='graded',
            title=f'Работа «{solution.work.title}» проверена',
            message=f'Оценка: {score}/{solution.work.max_score or 100}',
            link=f'/solution/{solution.id}/',
        )
        _notify_email(
            solution.student,
            f'Работа «{solution.work.title}» проверена',
            f'Оценка: {score}. Откройте решение для просмотра комментариев.',
        )

    messages.success(request, f'Оценка сохранена: {score}, статус: {new_status}')
    return redirect(request.POST.get('next', 'teacher_solutions'))


# ══════════════════════════════════════════════════════════════════════════════
# РЕЙТИНГ
# ══════════════════════════════════════════════════════════════════════════════

@login_required
def leaderboard(request):
    group_filter = request.GET.get('group', '')
    qs = UserProgress.objects.filter(
        user__is_active=True, user__is_staff=False
    ).select_related('user').order_by('-total_score', '-completed_works')

    all_groups = list(
        UserProgress.objects.filter(group__gt='').values_list('group', flat=True)
        .distinct().order_by('group')
    )  # TODO: add cache_page(60) decorator for performance
    if group_filter:
        qs = qs.filter(group=group_filter)

    leaders = list(qs[:50])
    my_rank = None
    for i, p in enumerate(leaders, 1):
        p.rank = i
        if p.user_id == request.user.id:
            my_rank = i

    return render(request, 'works/leaderboard.html', {
        'leaders': leaders,
        'my_rank': my_rank,
        'all_groups': all_groups,
        'group_filter': group_filter,
    })


# ══════════════════════════════════════════════════════════════════════════════
# ПОИСК ПО ТЕОРИИ
# ══════════════════════════════════════════════════════════════════════════════

@login_required
def theory_search(request):
    q = request.GET.get('q', '').strip()
    results = []
    if q and len(q) >= 2:
        from django.db.models import Q as DQ
        lessons = TheoryLesson.objects.filter(
            DQ(title__icontains=q) | DQ(content__icontains=q) | DQ(code_example__icontains=q)
        ).select_related('module')[:20]
        results = list(lessons)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = [{'id': l.id, 'title': l.title, 'module': l.module.title,
                 'url': f'/theory/lesson/{l.id}/'} for l in results]
        return JsonResponse({'results': data, 'count': len(data)})

    return render(request, 'works/theory_search.html', {'q': q, 'results': results})


# ── Смена пароля ────────────────────────────────────────────────────────────

@login_required
def change_password(request):
    """Простая страница смены пароля."""
    error = None
    success = False
    if request.method == 'POST':
        old_pw  = request.POST.get('old_password', '')
        new_pw1 = request.POST.get('new_password1', '')
        new_pw2 = request.POST.get('new_password2', '')
        if not request.user.check_password(old_pw):
            error = 'Неверный текущий пароль'
        elif len(new_pw1) < 6:
            error = 'Новый пароль должен быть не менее 6 символов'
        elif new_pw1 != new_pw2:
            error = 'Новые пароли не совпадают'
        else:
            request.user.set_password(new_pw1)
            request.user.save()
            # Переавторизуем чтобы сессия не сбросилась
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, request.user)
            success = True
    return render(request, 'works/change_password.html', {
        'error': error, 'success': success
    })


# ══════════════════════════════════════════════════════════════════════════════
# DEV AUTO-RELOAD
# ══════════════════════════════════════════════════════════════════════════════

def dev_reload(request):
    """Возвращает максимальный mtime отслеживаемых файлов. Только при DEBUG=True."""
    from django.conf import settings
    if not settings.DEBUG:
        return JsonResponse({'error': 'not available'}, status=403)

    latest = 0.0
    watch_dirs = [
        str(settings.BASE_DIR / 'works' / 'templates'),
        str(settings.BASE_DIR / 'works' / 'static'),
        str(settings.BASE_DIR / 'works'),
    ]
    skip_dirs = {'__pycache__', '.git', 'migrations', 'node_modules'}
    skip_exts = {'.pyc', '.pyo', '.sqlite3', '.log', '.lock'}

    for watch_dir in watch_dirs:
        if not os.path.exists(watch_dir):
            continue
        for root, dirs, files in os.walk(watch_dir):
            dirs[:] = [d for d in dirs if d not in skip_dirs]
            for fname in files:
                if os.path.splitext(fname)[1] in skip_exts:
                    continue
                try:
                    m = os.path.getmtime(os.path.join(root, fname))
                    if m > latest:
                        latest = m
                except OSError:
                    pass

    return JsonResponse({'mtime': latest})

# ══════════════════════════════════════════════════════════════════════════════
# СМЕНА ПАРОЛЯ
# ══════════════════════════════════════════════════════════════════════════════

@login_required
def change_password(request):
    """Смена пароля для студента и преподавателя."""
    error = None
    success = False

    if request.method == 'POST':
        old_pw  = request.POST.get('old_password', '')
        new_pw1 = request.POST.get('new_password1', '')
        new_pw2 = request.POST.get('new_password2', '')

        if not request.user.check_password(old_pw):
            error = 'Неверный текущий пароль'
        elif len(new_pw1) < 6:
            error = 'Новый пароль должен быть не менее 6 символов'
        elif new_pw1 != new_pw2:
            error = 'Пароли не совпадают'
        else:
            request.user.set_password(new_pw1)
            request.user.save()
            # Переавторизация после смены пароля
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, request.user)
            success = True
            messages.success(request, 'Пароль успешно изменён')

    return render(request, 'works/change_password.html', {
        'error': error,
        'success': success,
    })


# ══════════════════════════════════════════════════════════════════════════════
# ИСТОРИЯ РЕШЕНИЙ
# ══════════════════════════════════════════════════════════════════════════════

@login_required
def solution_history(request, work_id):
    """Все попытки сдачи конкретной работы студентом."""
    work = get_object_or_404(PracticalWork, id=work_id)
    # Студент видит только свои, преподаватель — все
    if request.user.is_staff:
        student_id = request.GET.get('student')
        if student_id:
            from django.contrib.auth.models import User as _User
            student = get_object_or_404(_User, id=student_id)
            attempts = Solution.objects.filter(work=work, student=student).order_by('-submitted_at')
        else:
            attempts = Solution.objects.filter(work=work).order_by('-submitted_at').select_related('student')
    else:
        student = request.user
        attempts = Solution.objects.filter(work=work, student=request.user).order_by('-submitted_at')

    return render(request, 'works/solution_history.html', {
        'work': work,
        'attempts': attempts,
        'is_teacher': request.user.is_staff,
    })


# ══════════════════════════════════════════════════════════════════════════════
# АНАЛИТИКА ПРЕПОДАВАТЕЛЯ
# ══════════════════════════════════════════════════════════════════════════════

@_staff_required
def teacher_analytics(request):
    """Дашборд с аналитикой успеваемости."""
    from django.utils import timezone
    from datetime import timedelta

    group_filter = request.GET.get('group', '')

    students_qs = UserProgress.objects.filter(
        user__is_active=True, user__is_staff=False
    ).select_related('user')
    if group_filter:
        students_qs = students_qs.filter(group=group_filter)

    all_groups = list(
        UserProgress.objects.filter(group__gt='').values_list('group', flat=True)
        .distinct().order_by('group')
    )

    works = PracticalWork.objects.filter(is_active=True).order_by('order')

    # Средний балл по каждой работе
    work_stats = []
    for w in works:
        sol_qs = Solution.objects.filter(work=w)
        if group_filter:
            sol_qs = sol_qs.filter(student__userprogress__group=group_filter)
        agg = sol_qs.aggregate(avg=Avg('score'), cnt=Count('id'), passed=Count('id', filter=Q(status='correct')))
        work_stats.append({
            'work': w,
            'avg_score':    round(agg['avg'] or 0, 1),
            'total':        agg['cnt'] or 0,
            'passed':       agg['passed'] or 0,
            'pass_rate':    round((agg['passed'] or 0) * 100 / agg['cnt']) if agg['cnt'] else 0,
        })

    # Самые сложные работы (наименьший pass_rate)
    hardest = sorted(work_stats, key=lambda x: x['pass_rate'])[:3]

    # Активность за последние 7 дней
    days = []
    for i in range(6, -1, -1):
        day = timezone.now().date() - timedelta(days=i)
        cnt = Solution.objects.filter(submitted_at__date=day).count()
        days.append({'date': day.strftime('%d.%m'), 'count': cnt})

    # Общая статистика
    total_students = students_qs.count()
    avg_score_all  = students_qs.aggregate(a=Avg('average_score'))['a'] or 0
    total_solutions = Solution.objects.count()
    passed_solutions = Solution.objects.filter(status='correct').count()

    # === Аналитика ошибок ===
    import json as _json
    from collections import Counter as _Counter

    # Топ-5 заданий с наибольшим кол-вом ошибочных решений
    error_works = (
        Solution.objects.filter(status='incorrect')
        .values('work__title', 'work_id')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')[:5]
    )

    # Частые ошибки из CodeCheck.errors
    error_counter = _Counter()
    for cc in CodeCheck.objects.exclude(errors=[]).order_by('-created_at')[:300]:
        try:
            errs = cc.errors if isinstance(cc.errors, list) else _json.loads(cc.errors or '[]')
            for e in errs:
                key = str(e)[:80].strip()
                if key:
                    error_counter[key] += 1
        except Exception:
            pass
    top_errors = error_counter.most_common(8)

    return render(request, 'works/teacher_analytics.html', {
        'work_stats':      work_stats,
        'hardest':         hardest,
        'days':            days,
        'total_students':  total_students,
        'avg_score_all':   round(avg_score_all, 1),
        'total_solutions': total_solutions,
        'passed_solutions': passed_solutions,
        'pass_rate_all':   round(passed_solutions * 100 / total_solutions) if total_solutions else 0,
        'all_groups':      all_groups,
        'group_filter':    group_filter,
        'error_works':     error_works,
        'top_errors':      top_errors,
    })


@login_required
@require_http_methods(["POST"])
def unlock_hint(request, hint_id):
    """Открыть подсказку за XP."""
    hint = get_object_or_404(WorkHint, id=hint_id)
    progress, _ = UserProgress.objects.get_or_create(user=request.user)
    already = UserHintUnlock.objects.filter(user=request.user, hint=hint).exists()
    if already:
        return JsonResponse({'text': hint.text, 'title': hint.title, 'already': True})
    if progress.total_score < hint.xp_cost:
        return JsonResponse({'error': f'Недостаточно очков. Нужно {hint.xp_cost}, у вас {progress.total_score}.'}, status=400)
    UserHintUnlock.objects.create(user=request.user, hint=hint)
    progress.total_score = max(0, progress.total_score - hint.xp_cost)
    progress.save(update_fields=['total_score'])
    return JsonResponse({'text': hint.text, 'title': hint.title, 'cost': hint.xp_cost})


@login_required
@require_http_methods(["POST"])
def request_deadline_extension(request, work_id):
    """Студент запрашивает продление дедлайна."""
    work = get_object_or_404(PracticalWork, id=work_id)
    reason = request.POST.get('reason', '').strip()
    if not reason:
        messages.error(request, 'Укажите причину запроса.')
        return redirect('work_detail', work_id=work_id)
    ext, created = DeadlineExtension.objects.get_or_create(
        user=request.user, work=work,
        defaults={'reason': reason, 'status': 'pending'}
    )
    if not created:
        ext.reason = reason
        ext.status = 'pending'
        ext.save(update_fields=['reason', 'status'])
    messages.success(request, 'Запрос отправлен преподавателю.')
    return redirect('work_detail', work_id=work_id)


@_staff_required
@require_http_methods(["POST"])
def approve_deadline_extension(request, ext_id):
    """Преподаватель одобряет или отклоняет продление."""
    ext = get_object_or_404(DeadlineExtension, id=ext_id)
    action = request.POST.get('action', 'approve')
    new_deadline_str = request.POST.get('new_deadline', '')
    if action == 'approve':
        ext.status = 'approved'
        if new_deadline_str:
            try:
                from django.utils.dateparse import parse_datetime
                ext.new_deadline = parse_datetime(new_deadline_str)
                if ext.new_deadline:
                    ext.work.deadline = ext.new_deadline
                    ext.work.save(update_fields=['deadline'])
            except Exception:
                pass
        Notification.send(
            user=ext.user, n_type='info',
            title=f'Продление дедлайна одобрено',
            message=f'Дедлайн по «{ext.work.title}» продлён.',
            link=f'/works/{ext.work.id}/',
        )
    else:
        ext.status = 'rejected'
        Notification.send(
            user=ext.user, n_type='info',
            title=f'Запрос на продление отклонён',
            message=f'Работа: «{ext.work.title}».',
        )
    ext.reviewed_at = timezone.now()
    ext.save(update_fields=['status', 'new_deadline', 'reviewed_at'])
    messages.success(request, 'Решение сохранено.')
    return redirect('teacher_solutions')


@_staff_required
def export_gradebook_excel(request):
    """Экспорт журнала успеваемости в Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    group_filter = request.GET.get('group', '')
    students_qs = UserProgress.objects.filter(
        user__is_active=True, user__is_staff=False
    ).select_related('user').order_by('user__last_name', 'user__first_name')
    if group_filter:
        students_qs = students_qs.filter(group=group_filter)

    works = list(PracticalWork.objects.filter(is_active=True).order_by('order'))
    students = list(students_qs)

    # Матрица оценок
    all_solutions = Solution.objects.filter(
        student__in=[s.user for s in students], work__in=works
    ).values('student_id', 'work_id', 'score', 'status').order_by('-score')
    matrix = {}
    for sol in all_solutions:
        key = (sol['student_id'], sol['work_id'])
        if key not in matrix:
            matrix[key] = sol

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Журнал успеваемости"

    # Стили
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="4472C4")
    pass_fill    = PatternFill("solid", fgColor="C6EFCE")
    partial_fill = PatternFill("solid", fgColor="FFEB9C")
    fail_fill    = PatternFill("solid", fgColor="FFC7CE")
    center       = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    # Заголовки
    headers = ["№", "Студент", "Группа"] + [f"ПР{w.order}" for w in works] + ["Среднее", "Сдано"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    ws.row_dimensions[1].height = 30
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 10

    # Данные
    for row_idx, prog in enumerate(students, 2):
        user = prog.user
        scores = []
        done = 0

        ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = center
        ws.cell(row=row_idx, column=2, value=f"{user.last_name} {user.first_name}")
        ws.cell(row=row_idx, column=3, value=prog.group or "—").alignment = center

        for col_idx, work in enumerate(works, 4):
            sol = matrix.get((user.id, work.id))
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center
            cell.border = thin_border
            if sol:
                cell.value = sol['score']
                if sol['status'] == 'correct':
                    cell.fill = pass_fill
                    done += 1
                    scores.append(sol['score'])
                elif sol['status'] == 'partially_correct':
                    cell.fill = partial_fill
                    scores.append(sol['score'])
                else:
                    cell.fill = fail_fill
            else:
                cell.value = "—"

        avg = round(sum(scores) / len(scores), 1) if scores else 0
        avg_cell = ws.cell(row=row_idx, column=len(works) + 4, value=avg)
        avg_cell.alignment = center
        avg_cell.font = Font(bold=True)
        ws.cell(row=row_idx, column=len(works) + 5, value=f"{done}/{len(works)}").alignment = center

    # Ширина колонок для работ
    for i in range(len(works)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 7

    # Легенда
    legend_row = len(students) + 3
    ws.cell(row=legend_row, column=1, value="Легенда:")
    for col, (label, fill) in enumerate([("Зачтено", pass_fill), ("Частично", partial_fill), ("Не зачтено", fail_fill)], 2):
        c = ws.cell(row=legend_row, column=col, value=label)
        c.fill = fill
        c.alignment = center

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"gradebook{'_' + group_filter if group_filter else ''}.xlsx"
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


# ══════════════════════════════════════════════════════════════════════════════
# ПРОГРЕСС-БЛОКИРОВКА МОДУЛЕЙ
# ══════════════════════════════════════════════════════════════════════════════

def get_module_unlock_status(user, modules):
    """
    Возвращает dict {module_id: is_unlocked}.
    Теория всегда читаема, тест модуля N требует прохождения теста N-1.
    """
    status = {}
    prev_passed = True  # первый модуль всегда открыт

    passed_quiz_modules = set(
        QuizAttempt.objects.filter(user=user, passed=True)
        .values_list('quiz__module_id', flat=True)
    )

    for module in sorted(modules, key=lambda m: m.order):
        if not module.requires_quiz_pass:
            status[module.id] = {'theory': True, 'quiz': True, 'locked': False}
        else:
            quiz_unlocked = prev_passed
            status[module.id] = {
                'theory': True,       # теория всегда открыта
                'quiz': quiz_unlocked,
                'locked': not quiz_unlocked,
            }
        prev_passed = module.id in passed_quiz_modules
    return status


# ══════════════════════════════════════════════════════════════════════════════
# КОНСТРУКТОР СХЕМ
# ══════════════════════════════════════════════════════════════════════════════

@login_required
def circuit_editor(request, work_id):
    """Открыть редактор схем для задания."""
    work = get_object_or_404(PracticalWork, id=work_id, is_active=True)
    draft = CircuitDraft.objects.filter(student=request.user, work=work).first()
    initial_json = draft.circuit_json if draft else 'null'
    return render(request, 'works/circuit_editor.html', {
        'work': work,
        'initial_json': initial_json,
        'readonly': False,
    })


@login_required
def circuit_editor_free(request):
    """Свободный редактор без привязки к заданию."""
    return render(request, 'works/circuit_editor.html', {
        'work': None,
        'initial_json': 'null',
        'readonly': False,
    })


@login_required
@require_http_methods(['POST'])
def circuit_save(request, work_id):
    """Автосохранение черновика схемы (AJAX)."""
    work = get_object_or_404(PracticalWork, id=work_id, is_active=True)
    try:
        body = request.body.decode('utf-8')
        json.loads(body)  # validate JSON
    except Exception:
        return JsonResponse({'status': 'error', 'msg': 'invalid json'}, status=400)
    draft, _ = CircuitDraft.objects.get_or_create(
        student=request.user, work=work,
        defaults={'circuit_json': body}
    )
    draft.circuit_json = body
    draft.save(update_fields=['circuit_json', 'updated_at'])
    return JsonResponse({'status': 'ok', 'saved_at': draft.updated_at.strftime('%H:%M:%S')})


@login_required
@require_http_methods(['POST'])
def circuit_submit(request, work_id):
    """Сдать схему как решение."""
    work = get_object_or_404(PracticalWork, id=work_id, is_active=True)
    circuit_json = request.POST.get('circuit_json', '{}')
    comment      = request.POST.get('comment', '')
    try:
        json.loads(circuit_json)
    except Exception:
        messages.error(request, 'Некорректные данные схемы.')
        return redirect('circuit_editor', work_id=work_id)

    sol = CircuitSolution.objects.create(
        student=request.user, work=work,
        circuit_json=circuit_json, comment=comment,
        status='submitted',
    )
    # Уведомление преподавателям
    for teacher in User.objects.filter(is_staff=True, is_active=True):
        Notification.objects.create(
            user=teacher,
            title=f'Новая схема от {request.user.last_name} {request.user.first_name}',
            message=f'Работа «{work.title}» — схема отправлена на проверку.',
            notification_type='submission',
        )
    messages.success(request, 'Схема отправлена на проверку!')
    return redirect('circuit_solution_detail', sol_id=sol.id)


@login_required
def circuit_solution_detail(request, sol_id):
    """Просмотр сданной схемы студентом."""
    sol = get_object_or_404(CircuitSolution, id=sol_id)
    if sol.student != request.user and not request.user.is_staff:
        return redirect('home')
    return render(request, 'works/circuit_editor.html', {
        'work': sol.work,
        'initial_json': sol.circuit_json,
        'readonly': True,
        'solution': sol,
    })


@login_required
def circuit_review(request, sol_id):
    """Преподаватель проверяет и оценивает схему."""
    if not request.user.is_staff:
        return redirect('home')
    sol = get_object_or_404(CircuitSolution, id=sol_id)
    if request.method == 'POST':
        status   = request.POST.get('status', 'reviewed')
        score    = int(request.POST.get('score', 0))
        comment  = request.POST.get('teacher_comment', '')
        sol.status = status
        sol.score  = min(score, sol.work.max_score)
        sol.teacher_comment = comment
        sol.reviewed_at = timezone.now()
        sol.save()
        Notification.objects.create(
            user=sol.student,
            title=f'Схема проверена: {sol.work.title}',
            message=f'Статус: {sol.get_status_display()}. Баллы: {sol.score}/{sol.work.max_score}.'
                    + (f'\nКомментарий: {comment}' if comment else ''),
            notification_type='grade',
        )
        messages.success(request, 'Оценка сохранена.')
        return redirect('circuit_solutions_list')
    return render(request, 'works/circuit_editor.html', {
        'work': sol.work,
        'initial_json': sol.circuit_json,
        'readonly': True,
        'solution': sol,
        'review_mode': True,
    })


@login_required
def circuit_solutions_list(request):
    """Список всех сданных схем для преподавателя."""
    if not request.user.is_staff:
        return redirect('home')
    sols = CircuitSolution.objects.select_related('student', 'work').order_by('-submitted_at')
    status_filter = request.GET.get('status', '')
    if status_filter:
        sols = sols.filter(status=status_filter)
    return render(request, 'works/circuit_solutions_list.html', {
        'solutions': sols[:100],
        'status_filter': status_filter,
    })
